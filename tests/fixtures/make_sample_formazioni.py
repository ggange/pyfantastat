"""Script to generate tests/fixtures/sample_formazioni.xlsx.

Run once:
    python tests/fixtures/make_sample_formazioni.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUT = Path(__file__).parent / "sample_formazioni.xlsx"

BLACK_FILL = PatternFill(fill_type="solid", fgColor="FF000000")
WHITE_FONT = Font(color="FFFFFFFF", bold=True)
BLACK_FONT = Font(color="FF000000")
GREEN_FONT = Font(color="FF008000")

wb = Workbook()
ws = wb.active
ws.title = "Formazioni 1 giornata"

def cell(r, c, value=None, fill=None, font=None):
    cell = ws.cell(row=r, column=c)
    if value is not None:
        cell.value = value
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    return cell

# Row 1: title
cell(1, 1, "Formazioni Test - Giornata 1")

# ---------------------------------------------------------------------------
# Left team: "ALPHA FC"  (cols A-E)
# Right team: "BETA SC"  (cols G-K)
# ---------------------------------------------------------------------------

# Row 4: team headers
cell(4, 1, "ALPHA FC",  fill=BLACK_FILL, font=WHITE_FONT)
cell(4, 7, "BETA SC",   fill=BLACK_FILL, font=WHITE_FONT)

# Row 5: formation
cell(5, 1, "433")
cell(5, 7, "442")

# Players: role=A/G, name=B/H, voto=D/J, fantavoto=E/K
players_left = [
    ("P", "Rossi",      6.5,  7.5),
    ("D", "Bianchi",    6.0,  6.0),
    ("D", "Verdi",      6.0,  6.5),
    ("D", "Neri",       5.5,  5.0),
    ("C", "Ferrari",    6.5,  7.0),
    ("C", "Romano",     7.0,  8.0),
    ("C", "Conti",      6.0,  6.0),
    ("A", "Russo",      7.5,  9.0),
    ("A", "Gallo",      5.5,  5.5),
    ("A", "Marini",     6.0,  6.0),
    # bench
    ("P", "Serra",      "-",  "-"),
    ("D", "Costa*",     "-",  "-"),
]

players_right = [
    ("P", "Bruno",      6.0,  6.5),
    ("D", "De Luca",    5.5,  5.5),
    ("D", "Fabbri",     6.5,  7.0),
    ("D", "Mancini",    6.0,  6.0),
    ("D", "Ferrara",    6.0,  6.0),
    ("C", "Ricci",      7.0,  8.5),
    ("C", "Gentile",    6.0,  6.5),
    ("A", "Amato",      5.5,  5.5),
    ("A", "Greco",      8.0, 12.0),
    # bench
    ("P", "Coppola",    "-",  "-"),
    ("C", "Monti",      "-",  "-"),
]

ROW_OFFSET = 6

for i, (role, name, voto, fv) in enumerate(players_left):
    r = ROW_OFFSET + i
    cell(r, 1, role)
    cell(r, 2, name.rstrip("*"))
    cell(r, 4, voto, font=GREEN_FONT)
    cell(r, 5, fv,   font=GREEN_FONT)

for i, (role, name, voto, fv) in enumerate(players_right):
    r = ROW_OFFSET + i
    cell(r, 7, role)
    cell(r, 8, name)
    cell(r, 10, voto, font=GREEN_FONT)
    cell(r, 11, fv,   font=GREEN_FONT)

# Totale rows (one row after last player)
last_row = ROW_OFFSET + max(len(players_left), len(players_right))

# Left totale: sum(fv) for played players = 7.5+6+6.5+7+8+6+9+5.5+6 = 61.5
cell(last_row, 1, "TOTALE: 61,50")
# Right totale: 6.5+5.5+7+6+6+8.5+6.5+5.5+12 = 63.5
cell(last_row + 1, 7, "TOTALE: 63,50")

# Inserita markers (end-of-block)
cell(last_row + 1, 1, "Inserita via web il 01-09-2024 20:00:00")
cell(last_row + 2, 7, "Inserita via app il 01-09-2024 21:00:00")

wb.save(OUT)
print(f"Written: {OUT}")
