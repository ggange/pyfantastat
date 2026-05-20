"""Parser for Fantacalcio "Formazioni" Excel files.

Each file covers one league matchday and contains all teams' lineups arranged
as side-by-side match blocks.  The parsed result exposes per-team totals and
per-player voto/fantavoto data.
"""

from __future__ import annotations

import re
import warnings
from pathlib import Path

from openpyxl import load_workbook

from pyfantastat.io.models import FormazioniMatchday, FormazioniPlayer, FormazioniTeam

# ---------------------------------------------------------------------------
# Column indices (1-based, as openpyxl uses)
# ---------------------------------------------------------------------------
_LEFT_ROLE_COL  = 1   # A
_LEFT_NAME_COL  = 2   # B
_LEFT_VOTO_COL  = 4   # D
_LEFT_FV_COL    = 5   # E
_LEFT_HDR_COL   = 1   # A  — where team-name header lives on the left side

_RIGHT_ROLE_COL = 7   # G
_RIGHT_NAME_COL = 8   # H
_RIGHT_VOTO_COL = 10  # J
_RIGHT_FV_COL   = 11  # K
_RIGHT_HDR_COL  = 7   # G

# Column in which the TOTALE row appears for each side
_LEFT_TOTALE_COL  = 1  # A
_RIGHT_TOTALE_COL = 7  # G

# Row-content prefixes that mark section boundaries, not real players
_SKIP_PREFIXES = (
    "panchina",
    "totale",
    "inserita",
    "recuperata",
    "modificatore",
)

_INSERITA_RE = re.compile(r"(inserita|recuperata)", re.IGNORECASE)
_GIORNATA_RE = re.compile(r"Giornata\s*(\d+)", re.IGNORECASE)
_TOTALE_RE   = re.compile(r"TOTALE:\s*(\d+)[,.](\d+)", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell_str(cell) -> str:
    v = cell.value
    return str(v).strip() if v is not None else ""


def _parse_score(val) -> float:
    """Parse a voto/fantavoto cell.  ``"-"`` or blank → 0.0."""
    if val is None:
        return 0.0
    s = str(val).strip()
    if s in ("-", ""):
        return 0.0
    try:
        return float(s.replace(",", "."))
    except ValueError:
        return 0.0


def _parse_totale(val) -> float | None:
    """Parse ``TOTALE: XX,XX`` → float, or ``None`` if not matched."""
    if val is None:
        return None
    m = _TOTALE_RE.search(str(val))
    return float(m.group(1) + "." + m.group(2)) if m else None


def _normalize_role(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip().upper()
    return s if s in ("P", "D", "C", "A") else None


def _is_team_header(cell) -> bool:
    """True for cells with black fill and white (or auto-white) font."""
    try:
        fill = cell.fill
        if fill is None or fill.fill_type != "solid":
            return False
        fg = fill.fgColor
        if fg is None:
            return False
        if fg.type == "rgb":
            if fg.rgb.upper() not in ("FF000000", "00000000"):
                return False
        elif fg.type == "theme":
            # Theme index 1 is black in the standard Office palette
            if fg.theme not in (0, 1):
                return False
        else:
            return False
        # Accept if font colour is white (or implicitly white because bg is black)
        fc = cell.font.color if cell.font else None
        if fc is not None and fc.type == "rgb":
            return fc.rgb.upper() in ("FFFFFFFF", "00FFFFFF")
        return True  # no explicit font colour recorded → treat as white on black
    except (AttributeError, TypeError):
        return False


def _should_skip(name: str) -> bool:
    low = name.lower()
    return name == "" or any(low.startswith(p) for p in _SKIP_PREFIXES)


# ---------------------------------------------------------------------------
# Side descriptor
# ---------------------------------------------------------------------------

class _Side:
    """Column offsets for the left or right half of a match block."""

    def __init__(self, totale_col: int) -> None:
        if totale_col == _LEFT_TOTALE_COL:
            self.role_col = _LEFT_ROLE_COL
            self.name_col = _LEFT_NAME_COL
            self.voto_col = _LEFT_VOTO_COL
            self.fv_col   = _LEFT_FV_COL
            self.hdr_col  = _LEFT_HDR_COL
        else:
            self.role_col = _RIGHT_ROLE_COL
            self.name_col = _RIGHT_NAME_COL
            self.voto_col = _RIGHT_VOTO_COL
            self.fv_col   = _RIGHT_FV_COL
            self.hdr_col  = _RIGHT_HDR_COL


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_formazioni_xlsx(path: str | Path) -> FormazioniMatchday:
    """Parse a Formazioni Excel file and return a :class:`FormazioniMatchday`.

    :param path: Path to the ``.xlsx`` file.
    :raises ValueError: If the giornata number cannot be extracted.
    """
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    max_row = ws.max_row or 0

    # ------------------------------------------------------------------
    # 1. Extract giornata from title row (first 5 rows, col A)
    # ------------------------------------------------------------------
    giornata: int | None = None
    for r in range(1, min(6, max_row + 1)):
        m = _GIORNATA_RE.search(_cell_str(ws.cell(r, 1)))
        if m:
            giornata = int(m.group(1))
            break
    if giornata is None:
        raise ValueError(
            f"Cannot extract giornata from '{path.name}'. "
            "Expected 'Giornata N' in the first 5 rows of column A."
        )

    # ------------------------------------------------------------------
    # 2. Locate "Inserita / Recuperata" marker rows
    # ------------------------------------------------------------------
    inserita_rows: list[int] = []
    for r in range(1, max_row + 1):
        v = _cell_str(ws.cell(r, 1)) or _cell_str(ws.cell(r, 7))
        if _INSERITA_RE.search(v):
            inserita_rows.append(r)

    # ------------------------------------------------------------------
    # 3. Find TOTALE cells near each marker (rows r, r-1, r-2)
    # ------------------------------------------------------------------
    seen: set[tuple[int, int]] = set()
    totale_entries: list[tuple[int, int, float]] = []  # (row, side_col, value)

    for ins_r in inserita_rows:
        for check_r in range(ins_r, max(0, ins_r - 4), -1):
            for col in (_LEFT_TOTALE_COL, _RIGHT_TOTALE_COL):
                if (check_r, col) in seen:
                    continue
                tot = _parse_totale(ws.cell(check_r, col).value)
                if tot is not None:
                    seen.add((check_r, col))
                    totale_entries.append((check_r, col, tot))

    # ------------------------------------------------------------------
    # 4. For each TOTALE, find team header and extract players
    # ------------------------------------------------------------------
    teams_seen: set[str] = set()
    team_list: list[FormazioniTeam] = []

    for tot_row, side_col, totale in totale_entries:
        side = _Side(side_col)

        # Search upward for white-on-black team-name header
        header_row: int | None = None
        team_name: str | None = None
        for r in range(tot_row - 1, max(0, tot_row - 51), -1):
            cell = ws.cell(r, side.hdr_col)
            raw = _cell_str(cell)
            if raw and _is_team_header(cell):
                header_row = r
                team_name = raw
                break

        if team_name is None or header_row is None:
            continue

        norm_key = team_name.strip().lower()
        if norm_key in teams_seen:
            continue
        teams_seen.add(norm_key)

        # Extract players between header+2 and totale row (exclusive)
        players: list[FormazioniPlayer] = []
        for r in range(header_row + 2, tot_row):
            raw_name = _cell_str(ws.cell(r, side.name_col)).rstrip("*").strip()
            if _should_skip(raw_name):
                continue
            players.append(FormazioniPlayer(
                name=raw_name,
                ruolo=_normalize_role(ws.cell(r, side.role_col).value),
                voto=_parse_score(ws.cell(r, side.voto_col).value),
                fantavoto=_parse_score(ws.cell(r, side.fv_col).value),
            ))

        team_list.append(FormazioniTeam(
            team_name=team_name,
            totale=totale,
            players=players,
        ))

    return FormazioniMatchday(giornata=giornata, teams=team_list)


def load_formazioni_dir(dir_path: str | Path) -> dict[int, FormazioniMatchday]:
    """Load all Formazioni Excel files from a directory.

    Matches files with the glob pattern ``Formazioni*.xlsx``.  Files that
    cannot be parsed emit a :class:`UserWarning` and are skipped.

    :returns: Mapping of 1-based giornata → :class:`FormazioniMatchday`.
    """
    dir_path = Path(dir_path)
    result: dict[int, FormazioniMatchday] = {}

    for xlsx_path in sorted(dir_path.glob("Formazioni*.xlsx")):
        try:
            md = load_formazioni_xlsx(xlsx_path)
        except Exception as exc:
            warnings.warn(f"Skipping '{xlsx_path.name}': {exc}", stacklevel=2)
            continue
        if md.giornata in result:
            warnings.warn(
                f"Duplicate giornata {md.giornata} — keeping first, "
                f"skipping '{xlsx_path.name}'.",
                stacklevel=2,
            )
            continue
        result[md.giornata] = md

    return result
