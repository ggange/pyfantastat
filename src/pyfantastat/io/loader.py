"""Excel calendar loader for Italian fantasy football (Fantacalcio) league files."""

import re
from collections import defaultdict
from pathlib import Path

from pyfantastat.io.models import CalendarData


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_float(value: object) -> float:
    text = _normalize_text(value).replace(",", ".")
    if not text or text == "-":
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _parse_giornata(text: str) -> int | None:
    match = re.search(r"(\d+)\s*ª?\s*Giornata", text, re.IGNORECASE)
    return int(match.group(1)) if match else None


def load_calendar_xlsx(path: str | Path) -> CalendarData:
    """Parse a Fantacalcio calendar workbook into a :class:`~pyfantastat.io.models.CalendarData`.

    The workbook is expected to follow the standard Italian fantasy football
    calendar layout:

    - Cell E1: ``"Calendario <league_name>"``
    - Columns A–F: odd-numbered league matchdays
    - Columns G–L: even-numbered league matchdays
    - Each match row: ``Team1 | Pts1 | Pts2 | Team2 | Result``

    :param path: Path to the ``.xlsx`` file.
    :returns: Parsed :class:`~pyfantastat.io.models.CalendarData`.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.worksheets[0]

    league_name: str | None = None
    cell_e1 = _normalize_text(sheet.cell(row=1, column=5).value)
    if cell_e1.lower().startswith("calendario "):
        league_name = cell_e1[len("Calendario "):].strip() or None

    team_names_set: set[str] = set()
    team_points: dict[str, list[float]] = defaultdict(list)
    calendar_by_matchday: dict[int, list[tuple[str, str]]] = defaultdict(list)
    serie_a_days: list[int] = []
    points_by_matchday: dict[int, list[tuple[float, float]]] = defaultdict(list)

    def process_match(
        row_idx: int,
        col_offset: int,
        league_day: int | None,
        serie_a_day: int | None,
    ) -> None:
        team_1 = _normalize_text(sheet.cell(row=row_idx, column=1 + col_offset).value)
        pts_1 = _parse_float(sheet.cell(row=row_idx, column=2 + col_offset).value)
        pts_2 = _parse_float(sheet.cell(row=row_idx, column=3 + col_offset).value)
        team_2 = _normalize_text(sheet.cell(row=row_idx, column=4 + col_offset).value)
        if not team_1 or not team_2:
            return
        team_names_set.add(team_1)
        team_names_set.add(team_2)
        team_points[team_1].append(pts_1)
        team_points[team_2].append(pts_2)
        if league_day is not None:
            calendar_by_matchday[league_day].append((team_1, team_2))
            points_by_matchday[league_day].append((pts_1, pts_2))
        if serie_a_day is not None:
            serie_a_days.append(serie_a_day)

    for row_idx in range(4, sheet.max_row + 1):
        left = _normalize_text(sheet.cell(row=row_idx, column=1).value)
        right = _normalize_text(sheet.cell(row=row_idx, column=7).value)

        if "Giornata lega" in left:
            league_day = _parse_giornata(left)
            serie_a_day = _parse_giornata(
                _normalize_text(sheet.cell(row=row_idx, column=4).value)
            )
            for next_row in range(row_idx + 1, min(sheet.max_row + 1, row_idx + 11)):
                if "Giornata lega" in _normalize_text(
                    sheet.cell(row=next_row, column=1).value
                ):
                    break
                process_match(next_row, 0, league_day, serie_a_day)

        if "Giornata lega" in right:
            league_day = _parse_giornata(right)
            serie_a_day = _parse_giornata(
                _normalize_text(sheet.cell(row=row_idx, column=10).value)
            )
            for next_row in range(row_idx + 1, min(sheet.max_row + 1, row_idx + 11)):
                if "Giornata lega" in _normalize_text(
                    sheet.cell(row=next_row, column=7).value
                ):
                    break
                process_match(next_row, 6, league_day, serie_a_day)

    team_names = sorted(team_names_set)
    team_name_to_index = {name: idx for idx, name in enumerate(team_names)}

    calendar: list[list[list[int]]] = []
    for league_day in sorted(calendar_by_matchday):
        day_matches: list[list[int]] = []
        for team_1, team_2 in calendar_by_matchday[league_day]:
            day_matches.append(
                [team_name_to_index[team_1], team_name_to_index[team_2]]
            )
        if day_matches:
            calendar.append(day_matches)

    current_matchday: int | None = None
    ordered_days = sorted(points_by_matchday)
    for idx, day in enumerate(ordered_days):
        if all(a == 0 and b == 0 for a, b in points_by_matchday[day]):
            current_matchday = ordered_days[idx - 1] if idx > 0 else None
            break
    if current_matchday is None and ordered_days:
        current_matchday = ordered_days[-1]

    match_count = max((len(pts) for pts in team_points.values()), default=0)

    return CalendarData(
        league_name=league_name,
        team_names=team_names,
        user_names=list(team_names),
        team_points=dict(team_points),
        match_count=match_count,
        calendar=calendar,
        current_matchday=current_matchday,
        first_matchday_serie_a=min(serie_a_days) if serie_a_days else None,
        last_matchday_serie_a=max(serie_a_days) if serie_a_days else None,
    )
